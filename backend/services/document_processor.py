import logging
import os
from typing import Any

logger = logging.getLogger("mochi.documents")

SUPPORTED_TYPES = {
    ".pdf": "pdf",
    ".docx": "docx",
    ".xlsx": "xlsx",
    ".csv": "csv",
    ".txt": "txt",
}


class DocumentProcessor:
    async def extract_text(self, file_path: str, file_type: str) -> str:
        ext = f".{file_type.lower()}"
        try:
            if ext == ".txt":
                return await self._read_txt(file_path)
            elif ext == ".csv":
                return await self._read_csv(file_path)
            elif ext == ".pdf":
                return await self._read_pdf(file_path)
            elif ext == ".docx":
                return await self._read_docx(file_path)
            elif ext == ".xlsx":
                return await self._read_xlsx(file_path)
            else:
                return ""
        except Exception as e:
            logger.error(f"Erro ao extrair texto de {file_path}: {e}")
            return ""

    async def _read_txt(self, path: str) -> str:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            return f.read()

    async def _read_csv(self, path: str) -> str:
        content = []
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            for line in f:
                content.append(line.strip())
        return "\n".join(content)

    async def _read_pdf(self, path: str) -> str:
        try:
            import PyPDF2
            text = []
            with open(path, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    text.append(page.extract_text() or "")
            return "\n".join(text)
        except ImportError:
            logger.warning("PyPDF2 nao instalado. Instale com: pip install PyPDF2")
            return "[PDF nao processado: PyPDF2 nao disponivel]"

    async def _read_docx(self, path: str) -> str:
        try:
            from docx import Document
            doc = Document(path)
            return "\n".join(p.text for p in doc.paragraphs)
        except ImportError:
            logger.warning("python-docx nao instalado. Instale com: pip install python-docx")
            return "[DOCX nao processado: python-docx nao disponivel]"

    async def _read_xlsx(self, path: str) -> str:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            text = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text.append(f"--- Planilha: {sheet} ---")
                for row in ws.iter_rows(values_only=True):
                    text.append(", ".join(str(c) if c is not None else "" for c in row))
            return "\n".join(text)
        except ImportError:
            logger.warning("openpyxl nao instalado. Instale com: pip install openpyxl")
            return "[XLSX nao processado: openpyxl nao disponivel]"

    async def summarize(self, text: str, max_length: int = 500) -> str:
        if not text.strip():
            return ""
        lines = text.split("\n")
        if len(text) <= max_length:
            return text
        return text[:max_length] + "..."

    async def extract_key_points(self, text: str) -> list[str]:
        if not text.strip():
            return []
        sentences = [s.strip() for s in text.replace("\n", ". ").split(".") if len(s.strip()) > 20]
        return sentences[:5]

    async def analyze_document(self, text: str) -> dict[str, Any]:
        return {
            "text_length": len(text),
            "line_count": len(text.split("\n")),
            "word_count": len(text.split()),
            "summary": await self.summarize(text),
            "key_points": await self.extract_key_points(text),
        }


document_processor = DocumentProcessor()
